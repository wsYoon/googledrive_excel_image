from httplib2 import Http
from oauth2client import file
import io
import os
import sys
from PyQt5 import QtCore,QtGui,QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import QSize, Qt
from PyQt5.QtWidgets import *
from PyQt5 import uic
from pdf2image import convert_from_path
#import win32api
import win32com
import win32com.client.gencache
import win32com.client as win32
import pythoncom
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
import datetime as dt
import math


def digit2txt(strNum):
    # 만 단위 자릿수
    tenThousandPos = 4
    # 억 단위 자릿수
    hundredMillionPos = 9
    txtDigit = ['', '십', '백', '천', '만', '억']
    txtNumber = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
    txtPoint = '쩜 '

    resultStr = ''
    digitCount = 0
    print(strNum)
    #자릿수 카운트
    for ch in strNum:
        # ',' 무시
        if ch == ',':
            continue
        #소숫점 까지
        elif ch == '.':
            break
        digitCount = digitCount + 1


    digitCount = digitCount-1
    index = 0

    while True:
        notShowDigit = False
        ch = strNum[index]
        #print(str(index) + ' ' + ch + ' ' +str(digitCount))
        # ',' 무시
        if ch == ',':
            index = index + 1
            if index >= len(strNum):
                break;
            continue

        if ch == '.':
            break
            #resultStr = resultStr + txtPoint
        else:
            #자릿수가 2자리이고 1이면 '일'은 표시 안함.
            # 단 '만' '억'에서는 표시 함
            if(digitCount > 1) and (digitCount != tenThousandPos) and  (digitCount != hundredMillionPos) and int(ch) == 1:
                resultStr = resultStr + ''
            elif int(ch) == 0:
                resultStr = resultStr + ''
                # 단 '만' '억'에서는 표시 함
                if (digitCount != tenThousandPos) and  (digitCount != hundredMillionPos):
                    notShowDigit = True
            else:
                resultStr = resultStr + txtNumber[int(ch)]


        # 1억 이상
        if digitCount > hundredMillionPos:
            if not notShowDigit:
                resultStr = resultStr + txtDigit[digitCount-hundredMillionPos]
        # 1만 이상
        elif digitCount > tenThousandPos:
            if not notShowDigit:
                resultStr = resultStr + txtDigit[digitCount-tenThousandPos]
        else:
            if not notShowDigit:
                resultStr = resultStr + txtDigit[digitCount]

        if digitCount <= 0:
            digitCount = 0
        else:
            digitCount = digitCount - 1
        index = index + 1
        if index >= len(strNum):
            break;
    return(resultStr)
    #print(resultStr)

form_class = uic.loadUiType("drive_excel_to_jpg.ui")[0]


#화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("CS이미지파일생성 프로그램")
        self.pushButton_Run.clicked.connect(self.Func_Run)          #실행버튼
        self.pushButton_Close.clicked.connect(self.Func_Close)   #위험수량 발주 버튼
        self.pushButton_Search.clicked.connect(self.Func_Search)       #마켓수수료 저장 버튼

    def Func_Run(self) :
        file_path = filename[0]

        #파일 수정을 위해서 열기
        recall_list = pd.read_excel(file_path,sheet_name='반품관리 최신 ',engine='openpyxl')
        input_list = pd.read_excel(file_path,sheet_name='송장번호',engine='openpyxl')
                    
        for songjang in input_list['송장번호']:
            songjang_str = str(songjang)
            if '.0' in songjang_str:
                songjang_str = songjang_str.replace('.0','')
            wb=openpyxl.load_workbook(file_path)
            ws = wb['롯데택배 합의서']
            ws.row_dimensions[2].height = 25
            print('송장번호',songjang)
            print(recall_list['보상요청가\n(상품가-상품가의20% +운임비1)'].loc[recall_list['운송장번호'] == songjang].item())
            amount = str(round(recall_list['보상요청가\n(상품가-상품가의20% +운임비1)'].loc[recall_list['운송장번호'] == songjang].item()))
            x = dt.datetime.now()
            ws.cell(row=4,column=4).value = songjang
            ws.cell(row=5,column=4).value = recall_list['상품명(노출상품명)'].loc[recall_list['운송장번호'] == songjang].item()
            ws.cell(row=7,column=9).value = digit2txt(amount) + ' 원(   ' + str(round(recall_list['보상요청가\n(상품가-상품가의20% +운임비1)'].loc[recall_list['운송장번호'] == songjang].item())) + '  )'
            #ws.cell(row=7,column=12).value = ' 원(      ' + str(round(recall_list['보상요청가\n(상품가-상품가의20% +운임비1)'].loc[recall_list['운송장번호'] == songjang].item())) + '  )'
            ws.cell(row=27,column=8).value = str(x.year) + ' .  ' + str(x.month) + '  .  ' + str(x.day) + ' .      구수룡   (인)'
            customer = recall_list['주문자명/수취인명'].loc[recall_list['운송장번호'] == songjang].item()
            customer = customer.replace('/','_')
            wb.save(file_path)

            # PDF 파일 경로
            #pdf_path = os.path.join(os.getcwd(), '구글드라이브다운로드', file_name1 + '_' + str(songjang) + '_' + str(customer) + '.pdf')
            pdf_path = os.path.join(os.getcwd(), '구글드라이브다운로드', songjang_str + '_' + str(customer) + '.pdf')

            # 엑셀 어플리케이션 생성
            excel_app = win32com.client.gencache.EnsureDispatch("Excel.Application",pythoncom.CoInitialize())
            #excel_app = win32.Dispatch("Excel.Application")
            xlLandscape = 1

            # 어플리케이션을 보이지 않게 설정
            excel_app.Visible = False

            # 엑셀 파일 열기
            workbook = excel_app.Workbooks.Open(file_path)

            # 두 번째 시트 선택
            worksheet = workbook.Worksheets(5)

            # 인쇄 설정 변경
            worksheet.PageSetup.Orientation = xlLandscape
            worksheet.PageSetup.FitToPagesTall = 1
            worksheet.PageSetup.FitToPagesWide = 1

            # PDF로 인쇄
            worksheet.ExportAsFixedFormat(0, pdf_path)

            # 엑셀 어플리케이션 종료
            # 변경사항 저장
            workbook.Save()

            # COM 객체 참조 해제
            worksheet = None
            workbook = None
            excel_app.Quit()
            excel_app = None

            # COM 객체 종료
            pythoncom.CoUninitialize()

            # pdf 파일에서 img 객체 리스트 반환해서 pages 변수로 받기
            pages = convert_from_path(pdf_path, poppler_path=os.getcwd()+r'\Release-23.01.0-0\poppler-23.01.0\Library\bin')

            # 이미지 파일 저장
            for i, page in enumerate(pages):
                page.save(f'구글드라이브다운로드/'+ songjang_str + '_' + str(customer) + '.jpg', 'JPEG')
                #page.save(f'구글드라이브다운로드/'+ file_name1 + str(songjang) + f'_{i}.jpg', 'JPEG')

    def Func_Search(self):
        global filename
        filename = QFileDialog.getOpenFileName(self, 'Open File')
        self.lineEdit_filename.setText(filename[0])

    def Func_Close(self):
        QApplication.quit()


if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #WindowClass의 인스턴스 생성
    myWindow = WindowClass() 

    #프로그램 화면을 보여주는 코드
    #myWindow.show()
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    app.exec_()